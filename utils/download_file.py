import os
from datetime import datetime

from django.conf import settings
from django.http import Http404, HttpResponse
import numpy as np
from openpyxl.styles import Border, Side
from openpyxl import load_workbook, Workbook

from geologiya.settings import BASE_DIR


def download(
        data, path, row_start_index, column_start_index, column_end_range, indexing
):
    file_name = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + ".xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, "excel", path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    file_path = os.path.join(file_path, file_name)
    index = [i for i in range(1, len(data) + 1)]
    try:
        if indexing:
            data = np.insert(data, 0, index, axis=0)
    except:
        pass

    shablon_path = os.path.join(BASE_DIR, "excel_template", path) + ".xlsx"
    book = load_workbook(shablon_path)
    sheet = book.active
    print(len(data))
    border = Border(top=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    left=Side(border_style='thin', color='FF000000'))
    for i in range(row_start_index, len(data) + row_start_index):
        print(row_start_index, i, row_start_index)
        for j in range(column_start_index, column_end_range + 1):
            print(column_start_index, j, column_end_range)
            x = sheet.cell(row=i, column=j)
            x.value = data[i - row_start_index][j - column_start_index]
            x.border = border

    book.save(file_path)

    content_type_value = "application/vnd.ms-excel"
    if os.path.exists(file_path):
        with open(file_path, "rb") as fh:
            response = HttpResponse(fh.read(), content_type=content_type_value)
            response["Content-Disposition"] = "inline; filename=" + file_name
        return response
    return Http404

